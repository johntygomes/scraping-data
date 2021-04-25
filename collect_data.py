import time
import openpyxl
import xlrd
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
#############################

number_of_diary = int(input("Enter Number Of Diary:: "))
number_of_diary = number_of_diary + 1    

###########################################################

driver=webdriver.Chrome()
driver.get("https://main.sci.gov.in/case-status")
time.sleep(7)
#pswd=driver.find_element_by_xpath("//input[@name='pass']")
#post=driver.find_element_by_xpath("//span[contains(text(),' on your mind, Guitarical?')]")
loc = 'collected_data_2.xlsx'
wb = openpyxl.load_workbook(filename=loc)
ws = wb.worksheets[0]

##########################################################                 
for m in range(2014,2022):
    try:
        sheet_year_title = "y"+str(m)
        ws_year = wb.create_sheet(sheet_year_title)
        ws_year = wb[sheet_year_title]
        new_m = m-2014
        year_string = "//option[@value='%s']"%str(m)
        select_year = driver.find_element_by_xpath("//select[@id='CaseDiaryYear']")        
        option_years = select_year.find_element_by_xpath(year_string)
        option_years.click()
        time.sleep(3)
        for k in range(1,number_of_diary):
            print(new_m,k)
            computed_row = (3*new_m) + k
            computed_row = abs(computed_row)
            print(computed_row)
            captcha_element  =  driver.find_element_by_xpath("//font[@color='red']")
            captcha_input    =  driver.find_element_by_xpath("//input[@name='ansCaptcha']")
            captcha_input.send_keys(captcha_element.text)
            time.sleep(2)

            case_diary_number = driver.find_element_by_xpath("//input[@id='CaseDiaryNumber']")
            temp = str(k)
            case_diary_number.clear()
            case_diary_number.send_keys(k)
            time.sleep(2)

            get_button = driver.find_element_by_xpath("//input[@id='getCaseDiary']")
            get_button.click()
            time.sleep(10)

            ####################################################################
            try:
                div = driver.find_element_by_xpath("//div[@id='accordion' and @class='panel-group']")
                table = div.find_element_by_tag_name("table")
                tr = table.find_elements_by_tag_name("tr")

                for i in range(0,len(tr)):#12
                    td = tr[i].find_elements_by_tag_name("td")
                    for j in range(0,len(td)):#2
                        my_cell_value = td[j].text
                        #print(my_cell_value)
                        if j==0:
                            #ws.cell(row=i+1, column=1, value=my_cell_value)
                            pass
                        else:                        
                            try:
                                ws.cell(row=computed_row, column=i+1, value=my_cell_value)
                                ws_year.cell(row=k+1, column=i+1, value=my_cell_value)
                            except Exception as e:
                                print(e)
            except Exception as e:
                print(e)
            wb.save(loc)
            time.sleep(3)
            print("##DONE##  ",m,k,i,j)
    ########################################################################
    except Exception as e:
        print(e,m,k,i,j)
    

